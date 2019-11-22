#coding=utf-8

import os
print(os.getcwd())
import sys
sys.path.append('C:/Users/18521/PycharmProjects/imuke')
import openpyxl

class HandleExcel(object):

    def  load_excel(self):
        wb=openpyxl.load_workbook('C:/Users/18521/PycharmProjects/imuke/Practice/imooc.xlsx')
        return wb

    def  get_sheetname(self,index=None):
        sheetname=self.load_excel().sheetnames
        if index == None:
            index=0
        sheetname=self.load_excel()[sheetname[index]]
        return sheetname

    def cell_value(self,row,column):
        data=self.get_sheetname().cell(row=row,column=column).value
        return data

    def get_rows(self):
        row_number=self.get_sheetname().max_row
        return row_number

    def get_row_data(self,row):
        row_list=[]
        for i in self.get_sheetname()[row]:
            row_list.append(i.value)
        return row_list

    def write_data(self,row,column,value):
        wb=self.load_excel()
        sh=wb.active
        sh.cell(row,column,value)
        wb.save('C:/Users/18521/PycharmProjects/imuke/Practice/imooc.xlsx')



ExcelData = HandleExcel()

if __name__=='__main__':
    e=HandleExcel()
    print(e.get_sheetname())
    print(e.cell_value(1,1))
    print(e.get_rows())
    print(e.get_row_data(1))

